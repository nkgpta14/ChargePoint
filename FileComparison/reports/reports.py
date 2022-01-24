from openpyxl.styles import Side, Alignment, Font, Border, PatternFill
from openpyxl.utils import get_column_letter
from data import Data
from library import ConfigReader


baseline_file = ConfigReader.readConfigData('Details', 'file1')
test_file = ConfigReader.readConfigData('Details', 'file2')
excel_report = ConfigReader.readConfigData('Details', 'Excel_Report')

def excel_formatting():

    try:
        workbook = Data.getWorkBook(excel_report)
        total_sheets = Data.getSheetName(excel_report)

        for sheet in total_sheets:
            worksheet = workbook[sheet]
            rows = Data.getRowCount(worksheet)
            columns = Data.getColumnCount(worksheet)

            for row_num in range(1, rows+1):
                for col_num in range(1, columns+1):
                    c = Data.readcell(worksheet, row_num, col_num)
                    thin = Side(border_style="thin", color="000000")
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.font = Font(name='Arial', bold=True, size=8, italic=False)
                    if row_num == 3 and col_num == 6:
                        if sheet == "Summary":
                            c.value = 'Compare Summary Report'
                        elif sheet == 'Details':
                            c.value = 'Detailed Report'
                        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        c.font = Font(name='Arial', bold=True, size=11, italic=False)
                        c.fill = PatternFill(start_color='82CAFF', end_color='82CAFF', fill_type='solid')
                        worksheet.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num+1, end_column=col_num+4)
                    if row_num == 6 and col_num >= 6:
                        c.fill = PatternFill(start_color='D1D0CE', end_color='D1D0CE', fill_type='solid')
                        length = (len(c.value) + 2)*1.2
                        worksheet.column_dimensions[get_column_letter(col_num)].width = length
                        c.font = Font(name='Arial Black', bold=False, size=9, italic=False)
                    if row_num >= 6 and col_num >= 6:
                        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if c.value == 'PASS':
                        c.fill = PatternFill(start_color='6CBB3C', end_color='6CBB3C', fill_type='solid')
                    elif c.value == 'FAIL':
                        c.fill = PatternFill(start_color='FF5050', end_color='FF5050', fill_type='solid')
                    if sheet == 'Details' and rows < 7:
                        Data.readcell(worksheet, 5, 6).value = 'Data Validation is Passed - No mismatches found'
                        Data.readcell(worksheet, 5, 6).fill = PatternFill(start_color='6CBB3C', end_color='6CBB3C', fill_type='solid')
                        Data.readcell(worksheet, 5, 6).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        worksheet.merge_cells(start_row=5, start_column=6, end_row=5, end_column=10)
        workbook.save(excel_report)
        print('Excel Report has been generated successfully')
    except Exception as e:
        print(f'excel_formatting - {e}')
        exit()
