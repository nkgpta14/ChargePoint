from openpyxl import load_workbook


def getWorkBook(path):

    workbook = load_workbook(path)
    return workbook

def getSheetName(path):

    total_sheets = []
    workbook = getWorkBook(path)
    sheets = workbook.sheetnames
    total_sheets = sheets
    return total_sheets


def getRowCount(sheet):

    return (sheet.max_row)

def getColumnCount(sheet):

    return (sheet.max_column)

def readcell(sheet, row_number,column_number):

    return (sheet.cell(row=row_number,column=column_number))
